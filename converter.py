# converter.py
import logging
import csv
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Set
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from utils import sanitize_filename, detect_csv_encoding

logger = logging.getLogger(__name__)

class ConversionCancelled(Exception):
    pass

class Converter:
    """
    执行物流出货数据格式转换的核心类。
    """
    # 常量定义
    BOX_CODE_COL = 0          # 第一列（0-based）
    SN_START_COL = 19         # 第20列（0-based）
    SN_END_COL = 258          # 第259列（0-based），共240列
    SN_COUNT = 240

    def __init__(self, input_file: Path, output_dir: Path,
                 duplicate_handling: str = 'suffix',
                 cancel_event=None):
        self.input_file = input_file
        self.output_dir = output_dir
        self.duplicate_handling = duplicate_handling  # 'skip', 'overwrite', 'suffix'
        self.cancel_event = cancel_event
        self.used_names: Set[str] = set()
        self.row_count = 0

    def _check_cancel(self):
        if self.cancel_event and self.cancel_event.is_set():
            raise ConversionCancelled("用户取消操作")

    def _read_xlsx(self):
        """读取 xlsx 文件，返回迭代器，每项为 (箱码, SN列表)"""
        wb = load_workbook(self.input_file, read_only=True, data_only=True)
        ws = wb.active
        # 获取标题行以确定列数（可选验证）
        # 使用迭代器读取，限制最大列到 SN_END_COL+1
        rows = ws.iter_rows(min_row=2, values_only=True,
                            max_col=self.SN_END_COL + 1)
        for row in rows:
            self._check_cancel()
            if not row or all(v is None for v in row):
                continue  # 跳过完全空行
            box_code = row[self.BOX_CODE_COL] if len(row) > self.BOX_CODE_COL else None
            # 提取 SN 列表，不足部分补 None
            sn_list = list(row[self.SN_START_COL:self.SN_END_COL + 1]) if len(row) > self.SN_START_COL else []
            sn_list += [None] * (self.SN_COUNT - len(sn_list))
            yield box_code, sn_list
        wb.close()

    def _read_csv(self):
        """读取 csv 文件，返回迭代器，每项为 (箱码, SN列表)"""
        encoding = detect_csv_encoding(self.input_file)
        with open(self.input_file, 'r', encoding=encoding, newline='') as f:
            reader = csv.reader(f)
            # 跳过标题行
            next(reader, None)
            for row in reader:
                self._check_cancel()
                if not row:
                    continue
                box_code = row[self.BOX_CODE_COL] if len(row) > self.BOX_CODE_COL else None
                sn_list = row[self.SN_START_COL:self.SN_END_COL + 1] if len(row) > self.SN_START_COL else []
                # 如果 SN 列不足，补 None
                if len(sn_list) < self.SN_COUNT:
                    sn_list += [None] * (self.SN_COUNT - len(sn_list))
                yield box_code, sn_list

    def _generate_output_filename(self, box_code, row_num: int) -> str:
        """生成清洗后的文件名（不含扩展名），处理重复和空箱码"""
        if box_code is None or str(box_code).strip() == '':
            base_name = f"row_{row_num}"
            logger.warning(f"第 {row_num} 行箱码为空，使用默认文件名 '{base_name}'")
        else:
            base_name = sanitize_filename(str(box_code).strip())
        # 检查是否重复
        if base_name in self.used_names:
            if self.duplicate_handling == 'skip':
                logger.warning(f"重复箱码 '{base_name}'，根据设置跳过该行")
                return None
            elif self.duplicate_handling == 'overwrite':
                logger.warning(f"重复箱码 '{base_name}'，覆盖已有文件")
                # 仍然记录名字，但会覆盖
            elif self.duplicate_handling == 'suffix':
                # 自动添加后缀
                counter = 1
                new_base = f"{base_name}_{counter}"
                while new_base in self.used_names:
                    counter += 1
                    new_base = f"{base_name}_{counter}"
                base_name = new_base
                logger.info(f"重复箱码，自动添加后缀为 '{base_name}'")
        self.used_names.add(base_name)
        return base_name

    def _write_single_xlsx(self, box_code, sn_list: List, row_num: int):
        """写入单个箱码的 xlsx 文件，返回输出文件路径"""
        base_name = self._generate_output_filename(box_code, row_num)
        if base_name is None:
            return None  # 跳过
        output_path = self.output_dir / f"{base_name}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "SN"
        ws.append(["SN"])
        for sn in sn_list:
            # 写入空字符串或实际值
            ws.append([sn if sn is not None else ""])
        wb.save(output_path)
        return output_path

    def run(self) -> Tuple[int, int]:
        """
        执行转换，返回 (成功数量, 总处理行数)
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if self.input_file.suffix.lower() == '.xlsx':
            row_iter = self._read_xlsx()
        elif self.input_file.suffix.lower() == '.csv':
            row_iter = self._read_csv()
        else:
            raise ValueError(f"不支持的文件格式: {self.input_file.suffix}")

        total_rows = 0
        success_count = 0
        for row_num, (box_code, sn_list) in enumerate(row_iter, start=2):
            self._check_cancel()
            total_rows += 1
            try:
                output_path = self._write_single_xlsx(box_code, sn_list, row_num)
                if output_path:
                    success_count += 1
                    logger.info(f"已生成: {output_path.name}")
            except Exception as e:
                logger.error(f"第 {row_num} 行处理失败: {e}", exc_info=True)
                # 根据需求，可以选择继续或中止，这里选择继续
                continue
        return success_count, total_rows
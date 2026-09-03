# gui.py
import os
import sys
import logging
import threading
from pathlib import Path
from PySide6.QtCore import Qt, QThread, Signal, QObject, Slot
from PySide6.QtGui import QDragEnterEvent, QDropEvent
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QFileDialog, QProgressBar, QTextEdit, QComboBox, QGroupBox,
    QRadioButton, QButtonGroup, QMessageBox, QApplication
)
from converter import Converter, ConversionCancelled
from utils import detect_csv_encoding

# 设置日志
logger = logging.getLogger(__name__)

class Worker(QObject):
    """后台转换工作类"""
    progress = Signal(int)          # 当前进度
    log = Signal(str)               # 日志消息
    finished = Signal(bool, str)    # 完成信号，参数为是否成功，错误信息（如失败）
    max_progress = Signal(int)      # 总行数

    def __init__(self, input_file: Path, output_dir: Path, duplicate_handling: str):
        super().__init__()
        self.input_file = input_file
        self.output_dir = output_dir
        self.duplicate_handling = duplicate_handling
        self.cancel_event = threading.Event()

    @Slot()
    def run(self):
        try:
            # 配置日志处理器，将日志发送到 GUI
            handler = GuiLogHandler(self.log)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
            converter = Converter(
                input_file=self.input_file,
                output_dir=self.output_dir,
                duplicate_handling=self.duplicate_handling,
                cancel_event=self.cancel_event
            )
            # 先计算总行数（用于进度条）
            total = self._count_rows()
            self.max_progress.emit(total)
            success, total_rows = converter.run()
            self.finished.emit(True, f"转换完成！成功生成 {success} 个文件，共处理 {total_rows} 行。")
        except ConversionCancelled:
            self.finished.emit(False, "转换已取消。")
        except Exception as e:
            logger.exception("转换过程中发生错误")
            self.finished.emit(False, f"转换失败：{str(e)}")
        finally:
            logger.removeHandler(handler)

    def _count_rows(self) -> int:
        """快速统计行数（不加载全部数据）"""
        if self.input_file.suffix.lower() == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(self.input_file, read_only=True)
            ws = wb.active
            # 使用 max_row 属性（近似值，但不精确），或者迭代计算
            count = ws.max_row - 1 if ws.max_row else 0
            wb.close()
            return max(0, count)
        elif self.input_file.suffix.lower() == '.csv':
            import csv
            encoding = detect_csv_encoding(self.input_file)
            with open(self.input_file, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                next(reader, None)  # 跳过标题
                count = sum(1 for _ in reader)
            return count
        else:
            return 0

    def cancel(self):
        self.cancel_event.set()

class GuiLogHandler(logging.Handler):
    """自定义日志处理器，将日志记录发送到 GUI 信号"""
    def __init__(self, signal):
        super().__init__()
        self.signal = signal

    def emit(self, record):
        msg = self.format(record)
        self.signal.emit(msg)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("物流出货数据格式转换工具")
        self.setMinimumSize(700, 500)
        self.setAcceptDrops(True)  # 启用拖拽
        self.worker_thread = None
        self.worker = None
        self._init_ui()
        self._setup_logging()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # 标题
        title_label = QLabel("物流出货数据格式转换工具")
        title_font = title_label.font()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # 文件选择区
        file_layout = QHBoxLayout()
        file_label = QLabel("输入文件:")
        self.input_edit = QLineEdit()
        self.input_edit.setPlaceholderText("请选择 .xlsx 或 .csv 文件，或拖拽文件到窗口")
        file_browse_btn = QPushButton("浏览...")
        file_browse_btn.clicked.connect(self.browse_input)
        file_layout.addWidget(file_label)
        file_layout.addWidget(self.input_edit, 1)
        file_layout.addWidget(file_browse_btn)
        main_layout.addLayout(file_layout)

        # 输出目录选择区
        out_layout = QHBoxLayout()
        out_label = QLabel("输出目录:")
        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText("默认在输入文件目录下创建 output 文件夹")
        out_browse_btn = QPushButton("浏览...")
        out_browse_btn.clicked.connect(self.browse_output)
        out_layout.addWidget(out_label)
        out_layout.addWidget(self.output_edit, 1)
        out_layout.addWidget(out_browse_btn)
        main_layout.addLayout(out_layout)

        # 重复箱码处理选项
        dup_group = QGroupBox("重复箱码处理")
        dup_layout = QHBoxLayout()
        self.dup_radio_suffix = QRadioButton("自动添加后缀")
        self.dup_radio_skip = QRadioButton("跳过")
        self.dup_radio_overwrite = QRadioButton("覆盖")
        self.dup_radio_suffix.setChecked(True)
        dup_btn_group = QButtonGroup(self)
        dup_btn_group.addButton(self.dup_radio_suffix)
        dup_btn_group.addButton(self.dup_radio_skip)
        dup_btn_group.addButton(self.dup_radio_overwrite)
        dup_layout.addWidget(self.dup_radio_suffix)
        dup_layout.addWidget(self.dup_radio_skip)
        dup_layout.addWidget(self.dup_radio_overwrite)
        dup_layout.addStretch()
        dup_group.setLayout(dup_layout)
        main_layout.addWidget(dup_group)

        # 按钮区
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始转换")
        self.start_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; padding: 8px 16px; }")
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.setEnabled(False)
        self.open_output_btn = QPushButton("打开输出目录")
        self.open_output_btn.setEnabled(False)
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.cancel_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.open_output_btn)
        main_layout.addLayout(btn_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        main_layout.addWidget(self.progress)

        # 日志区域
        log_label = QLabel("日志:")
        main_layout.addWidget(log_label)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        main_layout.addWidget(self.log_text, 1)  # 拉伸占据剩余空间

        # 信号连接
        self.start_btn.clicked.connect(self.start_conversion)
        self.cancel_btn.clicked.connect(self.cancel_conversion)
        self.open_output_btn.clicked.connect(self.open_output_folder)

    def _setup_logging(self):
        # 设置根日志记录器
        logging.basicConfig(level=logging.INFO,
                            format='%(asctime)s - %(levelname)s - %(message)s')
        # 将日志也输出到控制台
        console = logging.StreamHandler()
        console.setLevel(logging.INFO)
        logging.getLogger().addHandler(console)

    def browse_input(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择输入文件", "", "数据文件 (*.xlsx *.csv)")
        if file_path:
            self.input_edit.setText(file_path)
            # 自动填充输出目录（如果为空）
            if not self.output_edit.text().strip():
                input_path = Path(file_path)
                default_out = input_path.parent / "output"
                self.output_edit.setText(str(default_out))

    def browse_output(self):
        dir_path = QFileDialog.getExistingDirectory(
            self, "选择输出目录", self.output_edit.text() or "")
        if dir_path:
            self.output_edit.setText(dir_path)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        for url in event.mimeData().urls():
            path = Path(url.toLocalFile())
            if path.suffix.lower() in ['.xlsx', '.csv']:
                self.input_edit.setText(str(path))
                if not self.output_edit.text().strip():
                    default_out = path.parent / "output"
                    self.output_edit.setText(str(default_out))
                break

    def get_duplicate_handling(self) -> str:
        if self.dup_radio_skip.isChecked():
            return 'skip'
        elif self.dup_radio_overwrite.isChecked():
            return 'overwrite'
        else:
            return 'suffix'

    def start_conversion(self):
        input_str = self.input_edit.text().strip()
        output_str = self.output_edit.text().strip()
        if not input_str:
            QMessageBox.warning(self, "警告", "请选择输入文件！")
            return
        input_path = Path(input_str)
        if not input_path.exists():
            QMessageBox.critical(self, "错误", "输入文件不存在！")
            return
        if input_path.suffix.lower() not in ['.xlsx', '.csv']:
            QMessageBox.critical(self, "错误", "仅支持 .xlsx 或 .csv 格式！")
            return
        # 输出目录处理
        if not output_str:
            output_path = input_path.parent / "output"
        else:
            output_path = Path(output_str)
        try:
            output_path.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法创建输出目录：{e}")
            return

        # 禁用相关按钮
        self.start_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.open_output_btn.setEnabled(False)
        self.progress.setValue(0)
        self.log_text.clear()

        # 创建线程和 Worker
        self.worker_thread = QThread()
        self.worker = Worker(input_path, output_path, self.get_duplicate_handling())
        self.worker.moveToThread(self.worker_thread)

        # 连接信号
        self.worker_thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.update_progress)
        self.worker.max_progress.connect(self.set_progress_max)
        self.worker.log.connect(self.append_log)
        self.worker.finished.connect(self.on_finished)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)

        # 启动线程
        self.worker_thread.start()

    def cancel_conversion(self):
        if self.worker:
            self.worker.cancel()
            self.cancel_btn.setEnabled(False)

    def set_progress_max(self, max_val):
        self.progress.setRange(0, max_val if max_val > 0 else 1)
        self.progress.setValue(0)

    def update_progress(self, value):
        self.progress.setValue(value)

    def append_log(self, msg):
        self.log_text.append(msg)
        # 自动滚动到底部
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )

    def on_finished(self, success, message):
        self.start_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        if success:
            self.open_output_btn.setEnabled(True)
            QMessageBox.information(self, "完成", message)
        else:
            QMessageBox.warning(self, "转换结果", message)

    def open_output_folder(self):
        output_str = self.output_edit.text().strip()
        if not output_str:
            return
        output_path = Path(output_str)
        if output_path.exists():
            import subprocess
            import sys
            if sys.platform == 'win32':
                os.startfile(output_path)  # Windows
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', output_path])  # macOS
            else:
                subprocess.Popen(['xdg-open', output_path])  # Linux